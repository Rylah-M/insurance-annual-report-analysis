"""更新指标字典:新增分险种指标 + 修正非车非保证险保费描述(全表重建)。"""

from __future__ import annotations

import openpyxl


PATH = "agents/zd-agent0811/indicator/indicator_dictionary.xlsx"
SRC_PRIORITY = "经营情况分析|分部信息|主要险种经营信息|分险种信息"

ORIGINAL_IDS = {
    "F001", "F002", "F003", "F004", "F005", "F006", "F007", "F008", "F009", "F010",
    "B001", "B002", "B003", "B004", "B005", "B006", "B007", "B008", "B010",
    "B011", "B012", "B013", "B014", "B015", "B016", "B017",
    "R001", "R002", "R003",
}

# (显示名, 全称, 短称, 英文, 繁体全称, 繁体短称)
LINES = [
    ("车险", "车险", "机动车辆保险", "motor", "車險", "機動車輛保險"),
    ("非车险", "非车险", "非机动车辆保险", "non-motor", "非車險", "非機動車輛保險"),
    ("农险", "农业保险", "农险", "agricultural", "農業保險", "農險"),
    ("健康险", "健康险", "健康保险", "health", "健康險", "健康保險"),
    ("责任险", "责任保险", "责任险", "liability", "責任保險", "責任險"),
    ("意外险", "意外伤害保险", "意外险", "accident", "意外傷害保險", "意外險"),
    ("企财险", "企业财产保险", "企财险", "commercial property", "企業財產保險", "企財險"),
    ("保证险", "保证保险", "保证险", "guarantee", "保證保險", "保證險"),
    ("货运险", "货物运输保险", "货运险", "cargo transportation", "貨物運輸保險", "貨運險"),
    ("新能源车险", "新能源车险", "新能源车", "new energy vehicle", "新能源車險", "新能源車"),
]


def kw_metric(full: str, short: str, en: str, tf: str, ts: str, metric: str, metric_tw: str, en_metric: str) -> list[str]:
    kws = [
        f"{full}{metric}",
        f"{short}{metric}",
        f"{full}业务{metric}",
        f"{short}业务{metric}",
    ]
    if metric in ("保费", "保险服务收入", "保险服务费用"):
        kws += [f"{full}{metric}合计", f"{short}{metric}合计"]
    if metric == "保费":
        kws += [
            f"{full}{metric}收入", f"{short}{metric}收入",
            f"{full}原保险保费收入", f"{short}原保险保费收入",
            f"{full}原保费收入", f"{short}原保费收入",
        ]
    kws += [
        f"{tf}{metric_tw}", f"{ts}{metric_tw}",
        f"{tf}業務{metric_tw}", f"{ts}業務{metric_tw}",
        f"{en} {en_metric}",
    ]
    return kws


def join(items: list[str]) -> str:
    return "|".join(item for item in items if item)


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb["indicator_dictionary"]

    # 读取现有行(仅保留原始 29 行,其中 B004 之后单独修正)
    original = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in ORIGINAL_IDS:
            original.append(list(row))

    new_rows: list[list] = []

    # ---- 保险服务费用 (业务规模指标) B018-B027 ----
    for idx, (display, full, short, en, tf, ts) in enumerate(LINES, start=18):
        kws = kw_metric(full, short, en, tf, ts, "保险服务费用", "保險服務費用", "insurance service expenses")
        new_rows.append([
            f"B{idx:03d}", "业务规模指标", f"{display}保险服务费用",
            f"{full}保险服务费用|{short}保险服务费用|{en} insurance service expenses",
            join(kws),
            f"{full}业务确认的保险服务费用(IFRS 17 口径)。",
            "百万元", "number",
            "按年报分险种披露口径提取;未单独披露该险种保险服务费用则不提取。",
            SRC_PRIORITY,
        ])

    # ---- 保险服务收入 (业务规模指标, 补齐车险/非车险之外的险种) B028-B035 ----
    for idx, (display, full, short, en, tf, ts) in enumerate(LINES[2:], start=28):
        kws = kw_metric(full, short, en, tf, ts, "保险服务收入", "保險服務收入", "insurance service revenue")
        new_rows.append([
            f"B{idx:03d}", "业务规模指标", f"{display}保险服务收入",
            f"{full}保险服务收入|{short}保险服务收入|{en} insurance service revenue",
            join(kws),
            f"{full}业务确认的保险服务收入(IFRS 17 口径)。",
            "百万元", "number",
            "按年报分险种披露口径提取;未单独披露该险种保险服务收入则不提取。",
            SRC_PRIORITY,
        ])

    # ---- 新能源车险保费收入 B036 ----
    full, short, en, tf, ts = "新能源车险", "新能源车", "new energy vehicle", "新能源車險", "新能源車"
    kws = kw_metric(full, short, en, tf, ts, "保费", "保費", "premiums")
    new_rows.append([
        "B036", "业务规模指标", "新能源车险保费收入",
        "新能源车险原保险保费收入|新能源汽车保险保费|new energy vehicle premiums",
        join(kws + ["新能源汽车保险保费", "新能源车险原保费"]),
        "新能源车险(新能源汽车保险)业务的原保险保费收入。仅在公司披露新能源车险/新能源汽车保险数据时提取;未披露不得提取,严禁与整体车险指标混淆。",
        "百万元", "number",
        "按年报披露的'新能源车险/新能源汽车保险'保费提取;若未单独披露该险种保费,则不提取,不得用整体车险保费代替。",
        SRC_PRIORITY,
    ])

    # ---- 承保利润 (盈利能力指标) F011-F020 ----
    for idx, (display, full, short, en, tf, ts) in enumerate(LINES, start=11):
        kws = kw_metric(full, short, en, tf, ts, "承保利润", "承保利潤", "underwriting profit")
        if "新能源" in full:
            note = "仅在公司披露新能源车险/新能源汽车保险相关承保利润时提取;未披露不得提取,严禁与整体车险承保利润混淆。"
            calc = "按年报分险种披露口径提取;新能源车险须单独披露,否则不得提取。"
        else:
            note = ""
            calc = "按年报分险种披露口径提取;未披露则不提取。"
        definition = f"{full}业务承保端形成的利润或亏损。" + (f"{note}" if note else "")
        new_rows.append([
            f"F{idx:03d}", "盈利能力指标", f"{display}承保利润",
            f"{full}承保利润|{short}承保利润|{en} underwriting profit",
            join(kws + [f"{full}承保经营利润", f"{short}承保经营利润"]),
            definition,
            "百万元", "number", calc, SRC_PRIORITY,
        ])

    # ---- 综合成本率 (承保质量指标, 补齐车险/非车险之外的险种) F021-F028 ----
    for idx, (display, full, short, en, tf, ts) in enumerate(LINES[2:], start=21):
        kws = kw_metric(full, short, en, tf, ts, "综合成本率", "綜合成本率", "combined ratio")
        if "新能源" in full:
            note = "仅在公司披露新能源车险/新能源汽车保险综合成本率时提取;未披露不得提取,严禁与整体车险综合成本率混淆。"
            calc = "按年报分险种披露口径提取;新能源车险须单独披露,否则不得提取。"
        else:
            note = ""
            calc = "按年报分险种披露口径提取;未披露则不提取。"
        definition = f"{full}业务综合成本率(COR)。" + (f"{note}" if note else "")
        new_rows.append([
            f"F{idx:03d}", "承保质量指标", f"{display}综合成本率",
            f"{full}综合成本率|{short}综合成本率|{en} combined ratio",
            join(kws + [f"{full}COR", f"{short}COR", f"{full}承保综合成本率", f"{short}承保综合成本率"]),
            definition,
            "%", "percentage", calc, SRC_PRIORITY,
        ])

    # ---- 修正 B004 非车非保证险保费 ----
    for row in original:
        if row[0] == "B004":
            row[4] = (
                "非车非保证险|非车非保证保险|非车非保证险保费|非车非保证险保费收入"
                "|非车非保证保险保费|非车非保证保险保费收入"
                "|剔除保证保险后的非车险|不含保证保险的非车险|扣除保证保险的非车险|排除保证保险的非车险"
                "|non-auto non-guarantee premiums|非車非保證險|非車非保證保險"
                "|剔除保證保險後的非車險|不含保證保險的非車險"
            )
            row[5] = (
                "剔除保证保险后的非车险保费收入(=非车险保费收入-保证保险保费收入)。"
                "仅当年报明确披露'剔除/不含/扣除/排除保证保险'的非车险保费时提取。"
            )
            row[8] = (
                "若年报未单独披露保证保险保费,或未明确'剔除保证保险'口径,"
                "则该指标无对应数值,不得提取;严禁用非车险保费收入代替。"
            )

    # 全表重建:表头 + 原始行 + 新增行
    ws.delete_rows(2, ws.max_row - 1)
    for row in original + new_rows:
        ws.append(row)
    wb.save(PATH)
    print(f"重建完成: 原始 {len(original)} 行 + 新增 {len(new_rows)} 行 = {len(original) + len(new_rows)} 行")


if __name__ == "__main__":
    main()
